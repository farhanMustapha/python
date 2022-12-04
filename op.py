#https://www.youtube.com/watch?v=frceeS7w5eQ

from openpyxl import Workbook,load_workbook

wb=Workbook()
ws=wb.active

ws.title="New Title"
another_ws=wb.create_sheet("MySheet")
ws.cell(column=3,row=1,value=10)
ws['A3']=20
ws.append(["111","2222",22,44])
wb.save("MyWorkBook.xlsx")